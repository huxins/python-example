from openpyxl import load_workbook

def xstr(s):
    if s is None:
        return ''
    return str(s)

workbook = load_workbook(filename='转换为文本格式.xlsx')
data_list = []
text_list = []
for sheetname in workbook.sheetnames:
    worksheet = workbook[sheetname]
    for row in range(1,worksheet.max_row+1):
        data_list_row = []
        for column in range(1,worksheet.max_column+1):
            data_list_row.append(worksheet.cell(row=row, column=column).value)
        data_list.append(data_list_row)
for x in data_list:
    text_list.append(','.join(map(xstr,x)))

with open('somefile.txt', 'wt') as f:
    f.write('\n'.join(text_list))
